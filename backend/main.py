import os
from fastapi import FastAPI, HTTPException, File, UploadFile, Form
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from supabase import create_client, Client
from dotenv import load_dotenv
from pydantic import BaseModel
from typing import List, Optional
import math
# import pandas as pd (Removed for lightness)
from datetime import datetime
import io

# Try to load .env file (for local development)
if os.path.exists("../.env"):
    load_dotenv(dotenv_path="../.env")
else:
    # Render/Production will have env vars set directly
    load_dotenv()

app = FastAPI(title="Academic Command Center API")

# Setup CORS to allow Next.js frontend to talk to FastAPI
# For production, we allow all origins or specify the Vercel URL
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Supabase client
SUPABASE_URL = os.getenv("NEXT_PUBLIC_SUPABASE_URL") or os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("NEXT_PUBLIC_SUPABASE_ANON_KEY") or os.getenv("SUPABASE_ANON_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    pass

supabase: Client | None = None
if SUPABASE_URL and SUPABASE_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
else:
    pass

@app.get("/api/health")
def health_check():
    status = "healthy"
    supabase_status = "connected" if supabase else "disconnected"
    return {"status": status, "supabase": supabase_status, "message": "API backend is running!"}

@app.get("/api/test-supabase")
def test_supabase():
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    return {"status": "success", "message": "Supabase client is successfully configured."}

class SubjectCreate(BaseModel):
    name: str
    code: str
    semester: str

class TaskCreate(BaseModel):
    title: str
    deadline: str
    status: str
    subject_id: str

class SessionCreate(BaseModel):
    class_id: str
    lat: float
    lng: float

class AttendanceSubmit(BaseModel):
    session_id: str
    mssv: str
    lat: float
    lng: float
    device_id: str

def calculate_distance(lat1, lon1, lat2, lon2):
    # Earth radius in meters
    R = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    
    a = math.sin(dphi / 2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

@app.get("/api/classes")
def get_classes():
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        # Join with subjects table to get the subject name (Normalization)
        response = supabase.table("classes").select("*, subjects(name)").execute()
        return response.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/classes/{class_id}")
def delete_class(class_id: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        res = supabase.table("classes").delete().eq("id", class_id).execute()
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/subjects")
def get_subjects():
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        response = supabase.table("subjects").select("*").execute()
        return response.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/subjects")
def create_subject(subject: SubjectCreate):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        data, count = supabase.table("subjects").insert({
            "name": subject.name,
            "code": subject.code,
            "semester": subject.semester
        }).execute()
        return data[1][0] if len(data) > 1 and len(data[1]) > 0 else data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/tasks")
def get_tasks():
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        # Join with subjects to get the subject name
        response = supabase.table("tasks").select("*, subjects(name)").execute()
        return response.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/tasks")
def create_task(task: TaskCreate):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        data, count = supabase.table("tasks").insert({
            "title": task.title,
            "deadline": task.deadline,
            "status": task.status,
            "subject_id": task.subject_id
        }).execute()
        return data[1][0] if len(data) > 1 and len(data[1]) > 0 else data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/api/tasks/{task_id}")
def update_task_status(task_id: str, status: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        data, count = supabase.table("tasks").update({"status": status}).eq("id", task_id).execute()
        return data[1][0] if len(data) > 1 and len(data[1]) > 0 else data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents")
def get_documents():
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        # Fetch documents joined with subject name
        response = supabase.table("documents").select("*, subjects(name)").execute()
        return response.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/documents/upload")
async def upload_document(
    file: UploadFile = File(...),
    name: str = Form(...),
    subject_id: str = Form(...)
):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    
    try:
        # 1. Read file content
        file_content = await file.read()
        file_ext = os.path.splitext(file.filename)[1]
        file_path = f"{subject_id}/{os.urandom(8).hex()}_{file.filename}"
        
        # 2. Upload to Supabase Storage
        # Note: Bucket must be created manually or we handle it if possible.
        # supabase.storage.create_bucket('academic-docs') might fail if no permissions.
        storage_response = supabase.storage.from_("academic-docs").upload(
            path=file_path,
            file=file_content,
            file_options={"content-type": file.content_type}
        )
        
        # 3. Get Public URL
        public_url = supabase.storage.from_("academic-docs").get_public_url(file_path)
        
        # 4. Save metadata to database
        doc_data = {
            "name": name,
            "file_url": public_url,
            "file_path": file_path,
            "subject_id": subject_id,
            "file_type": file.content_type or file_ext
        }
        
        db_response = supabase.table("documents").insert(doc_data).execute()
        return db_response.data[0] if db_response.data else doc_data
        
    except Exception as e:
        
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/dashboard/summary")
def get_dashboard_summary():
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        # 1. Get counts
        subjects_res = supabase.table("subjects").select("id", count="exact").execute()
        tasks_res = supabase.table("tasks").select("id", count="exact").neq("status", "Hoàn thành").execute()
        docs_res = supabase.table("documents").select("id", count="exact").execute()
        
        # 2. Get recent tasks (Top 5)
        recent_tasks = supabase.table("tasks").select("*, subjects(name)").neq("status", "Hoàn thành").order("deadline").limit(5).execute()
        
        return {
            "stats": {
                "subjects": subjects_res.count if subjects_res.count is not None else 0,
                "tasks_pending": tasks_res.count if tasks_res.count is not None else 0,
                "documents": docs_res.count if docs_res.count is not None else 0
            },
            "recent_tasks": recent_tasks.data or []
        }
    except Exception as e:
        
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/attendance/session")
def create_attendance_session(session: SessionCreate):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        target_class_id = session.class_id
        
        # Nếu class_id gửi lên không phải UUID (mà là mã lớp như 225NMG03)
        # thì ta đi tìm ID thực sự của nó trong bảng classes
        if len(str(target_class_id)) < 20: # Heuristic check for non-UUID
            class_res = supabase.table("classes").select("id").eq("ma_lop", str(target_class_id).strip()).execute()
            if class_res.data:
                target_class_id = class_res.data[0]["id"]
            else:
                # Nếu chưa có lớp này trong bảng classes, tạo mới luôn
                new_class = supabase.table("classes").insert({
                    "ma_lop": str(target_class_id).strip(),
                    "ten_mon": f"Lớp {target_class_id}"
                }).execute()
                if new_class.data:
                    target_class_id = new_class.data[0]["id"]

        response = supabase.table("attendance_sessions").insert({
            "class_id": target_class_id,
            "teacher_lat": session.lat,
            "teacher_lng": session.lng,
            "status": "active"
        }).execute()
        return response.data[0] if response.data else {"status": "error"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/attendance/submit")
def submit_attendance(submission: AttendanceSubmit):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        # 1. Get session info to get teacher's location
        session_res = supabase.table("attendance_sessions").select("*").eq("id", submission.session_id).single().execute()
        if not session_res.data:
            raise HTTPException(status_code=404, detail="Buổi điểm danh không tồn tại.")
        
        session = session_res.data
        if session["status"] != "active":
            raise HTTPException(status_code=400, detail="Buổi điểm danh này đã kết thúc.")

        # --- Kiểm tra Chống gian lận (Device Locking) ---
        student_res = supabase.table("students").select("*").eq("mssv", submission.mssv).execute()
        if not student_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy thông tin sinh viên.")
        student = student_res.data[0]

        current_device_id = student.get("current_device_id")
        
        if not current_device_id:
            # Bước B: Lần đầu tiên điểm danh, cập nhật device_id
            supabase.table("students").update({"current_device_id": submission.device_id}).eq("mssv", submission.mssv).execute()
        else:
            # Bước C: So sánh device_id
            if current_device_id != submission.device_id:
                raise HTTPException(status_code=400, detail="Thiết bị này không khớp với thiết bị bạn đã đăng ký. Vui lòng liên hệ giảng viên để reset.")

        # Bước D: Kiểm tra 1 máy không được điểm danh cho nhiều người trong cùng session
        attended_res = supabase.table("attendance_records").select("mssv").eq("session_id", submission.session_id).execute()
        attended_mssvs = [r["mssv"] for r in attended_res.data if r["mssv"] != submission.mssv]
        
        if attended_mssvs:
            duplicate_device_res = supabase.table("students").select("mssv").eq("current_device_id", submission.device_id).in_("mssv", attended_mssvs).execute()
            if duplicate_device_res.data:
                raise HTTPException(status_code=400, detail="Thiết bị này đã được sử dụng để điểm danh cho một sinh viên khác trong buổi học này.")
        # -----------------------------------------------

        # 2. Calculate distance
        dist = calculate_distance(
            session["teacher_lat"], session["teacher_lng"],
            submission.lat, submission.lng
        )

        # 3. Check distance (<= 100m)
        if dist > 100:
            raise HTTPException(status_code=400, detail=f"Bạn đang ở quá xa vị trí điểm danh ({dist:.1f}m).")

        # 4. Save record
        record_res = supabase.table("attendance_records").insert({
            "session_id": submission.session_id,
            "mssv": submission.mssv,
            "student_lat": submission.lat,
            "student_lng": submission.lng,
            "distance": dist,
            "device_id": submission.device_id 
        }).execute()
        
        return {
            "status": "success",
            "message": "Điểm danh thành công!",
            "distance": dist,
            "data": record_res.data[0] if record_res.data else None
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/api/attendance/session/{session_id}")
def update_attendance_session_status(session_id: str, status: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        res = supabase.table("attendance_sessions").update({"status": status}).eq("id", session_id).execute()
        return res.data[0] if res.data else {"status": "error"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

import openpyxl
import openpyxl.styles



@app.post("/api/admin/sync-students")
async def sync_students_from_excel(file: UploadFile = File(...)):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase not initialized.")

    try:
        # Load workbook with data_only=True
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        sheet = wb.active

        # 1. Metadata Detection (Header & Class Info)
        start_row = 1
        class_code = None
        subject_name = None
        
        for r in range(1, 16):
            for c in range(1, 10):
                val = str(sheet.cell(row=r, column=c).value or "").strip()
                # Class Code (7+ chars with digits)
                if not class_code and len(val) >= 7 and any(char.isdigit() for char in val):
                    if r < 5: class_code = val
                # Subject Name
                if "Tên học phần" in val:
                    subject_name = str(sheet.cell(row=r, column=c+2).value or "").strip()
                    if " - Số TC:" in subject_name:
                        subject_name = subject_name.split(" - Số TC:")[0].strip()
                # Data Header
                if "Mã SV" in val or "MSSV" in val:
                    start_row = r + 1

        if not class_code:
            class_code = sheet.cell(row=1, column=5).value or "UNKNOWN_CLASS"

        # 2. Scientific Data Cleaning (Full Scan Approach)
        processed_students = []
        stats = {"total_rows_found": 0, "successfully_synced": 0, "errors": 0}
        
        # Quét toàn bộ từ dòng bắt đầu đến dòng cuối cùng của Sheet
        print(f"DEBUG: Bat dau quet tu dong {start_row} den {sheet.max_row}")
        
        for row_idx in range(start_row, sheet.max_row + 1):
            stats["total_rows_found"] += 1
            
            # Step 2.1: Extraction & Extreme Cleaning
            mssv_raw = sheet.cell(row=row_idx, column=2).value
            # Chi lay cac ky tu la so (Giai quyet loi 2415.0 -> 2415)
            mssv = "".join(filter(str.isdigit, str(mssv_raw or "")))
            
            first_name = str(sheet.cell(row=row_idx, column=3).value or "").strip()
            last_name = str(sheet.cell(row=row_idx, column=4).value or "").strip()
            full_name = f"{first_name} {last_name}".strip()
            
            if mssv or full_name:
                print(f"DEBUG: Dang doc dong {row_idx}, MSSV: {mssv}, Ten: {full_name}")

            # Step 2.2: Scientific Validation (8-15 characters for MSSV)
            if len(mssv) < 8 or len(mssv) > 15 or not full_name or len(full_name) < 2:
                if mssv or full_name:
                    print(f"DEBUG: Bo qua dong {row_idx} do du lieu khong hop le (MSSV: '{mssv}' dai {len(mssv)})")
                stats["errors"] += 1
                continue
            
            email = str(sheet.cell(row=row_idx, column=6).value or f"{mssv}@student.edu.vn").strip()
            processed_students.append({"mssv": mssv, "name": full_name, "email": email})
            print(f"DEBUG: -> NHAT DUOC: {mssv} - {full_name}")

        print(f"DEBUG: Ket thuc quet. Tong so sinh vien nhat duoc: {len(processed_students)}")

        if not processed_students:
            return {**stats, "detail": "Không có dữ liệu hợp lệ (Kiem tra lai MSSV dai 8-12 ky tu)."}

        # 3. Database Sync (Relational Strategy)
        class_code_str = str(class_code).strip()
        semester_str = class_code_str[:3] if class_code_str[:3].isdigit() else "HK"
        final_subject_name = subject_name if subject_name else "Chưa xác định"
        
        # 3.1: Upsert Subject & Get subject_id
        subject_id = None
        try:
            import re
            subj_code_match = re.search(r'[A-Za-z]+', class_code_str)
            subj_code = subj_code_match.group(0) if subj_code_match else class_code_str
            sub_res = supabase.table("subjects").upsert({
                "name": final_subject_name,
                "code": subj_code,
                "semester": semester_str
            }, on_conflict="name").execute()
            
            if sub_res.data:
                subject_id = sub_res.data[0]["id"]
        except Exception as e: 
            print(f"Subject Sync Error: {e}")

        # 3.2: Upsert Class with BOTH subject_id and ten_mon (to satisfy Not-Null constraint)
        class_res = supabase.table("classes").select("id").match({"ma_lop": class_code_str, "semester": semester_str}).execute()
        if not class_res.data:
            class_obj = supabase.table("classes").insert({
                "ma_lop": class_code_str,
                "subject_id": subject_id,
                "ten_mon": final_subject_name, # Satisfy legacy constraint
                "semester": semester_str
            }).execute()
            class_id = class_obj.data[0]["id"]
        else:
            class_id = class_res.data[0]["id"]
            # Update both for consistency
            supabase.table("classes").update({
                "subject_id": subject_id,
                "ten_mon": final_subject_name
            }).eq("id", class_id).execute()

        # 3.3: Batch Upsert Students
        for i in range(0, len(processed_students), 100):
            batch = processed_students[i:i+100]
            supabase.table("students").upsert(batch, on_conflict="mssv").execute()
            
        # 3.4: Batch Upsert Junction (Enrollment)
        junction_data = [{"class_id": class_id, "mssv": s["mssv"]} for s in processed_students]
        for i in range(0, len(junction_data), 100):
            batch = junction_data[i:i+100]
            supabase.table("class_students").upsert(batch, on_conflict="class_id, mssv").execute()

        stats["successfully_synced"] = len(processed_students)
        return stats

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Scientific Sync Error: {str(e)}")

@app.post("/api/attendance/session/{session_id}/finalize")
async def finalize_attendance(session_id: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    
    try:
        # 1. Mark session as inactive
        session_res = supabase.table("attendance_sessions").update({"status": "inactive"}).eq("id", session_id).execute()
        if not session_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiên điểm danh.")
        
        session_info = session_res.data[0]
        class_id = session_info["class_id"]
        
        # Get class info
        class_res = supabase.table("classes").select("*").eq("id", class_id).single().execute()
        class_name = class_res.data["ten_mon"] if class_res.data else "Lớp học"

        # 2. Get all students for this class from class_students junction table
        students_res = supabase.table("class_students").select("mssv").eq("class_id", class_id).execute()
        all_students_mssvs = [s["mssv"] for s in (students_res.data or [])]

        # 3. Get all present students for this session
        present_res = supabase.table("attendance_records").select("mssv").eq("session_id", session_id).execute()
        present_mssvs = {r["mssv"] for r in present_res.data}

        # 4. Find absent students
        absent_count = len(all_students_mssvs) - len(present_mssvs)
        if absent_count < 0: absent_count = 0 # Safety check

        return {
            "status": "success",
            "present_count": len(present_mssvs),
            "absent_count": absent_count
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/api/students/{mssv}/reset-device")
def reset_student_device(mssv: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        res = supabase.table("students").update({"current_device_id": None}).eq("mssv", mssv).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên.")
        return {"status": "success", "message": "Đã reset thiết bị thành công"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/attendance/session/{session_id}")
async def get_session_status(session_id: str):
    try:
        # 1. Lấy thông tin phiên
        session_res = supabase.table("attendance_sessions").select("*, classes(*)").eq("id", session_id).single().execute()
        if not session_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiên.")
            
        session_info = session_res.data
        class_id = session_info["class_id"]
        
        # 2. Lấy TOÀN BỘ danh sách sinh viên của lớp này
        students_res = supabase.table("class_students").select("mssv, students(name)").eq("class_id", class_id).execute()
        all_students = students_res.data or []
        
        # 3. Lấy những người ĐÃ điểm danh trong phiên này
        records_res = supabase.table("attendance_records").select("mssv, distance").eq("session_id", session_id).execute()
        present_records = {r["mssv"]: r for r in (records_res.data or [])}
        
        # 4. Phân loại
        present_list = []
        absent_list = []
        
        for item in all_students:
            mssv = item["mssv"]
            name = item["students"]["name"] if item.get("students") else "N/A"
            
            if mssv in present_records:
                present_list.append({
                    "mssv": mssv,
                    "name": name,
                    "status": "present",
                    "distance": present_records[mssv].get("distance")
                })
            else:
                absent_list.append({
                    "mssv": mssv,
                    "name": name,
                    "status": "absent"
                })
                
        return {
            "session": session_info,
            "present": present_list,
            "absent": absent_list
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/attendance/sessions/{session_id}/export-absentees")
def export_absentees(session_id: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        # 1. Get session info to get class_id
        session_res = supabase.table("attendance_sessions").select("class_id").eq("id", session_id).single().execute()
        if not session_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiên điểm danh")
            
        class_id = session_res.data["class_id"]
        
        # 2. Get class info
        class_res = supabase.table("classes").select("ma_lop", "ten_mon").eq("id", class_id).single().execute()
        if not class_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy lớp học")
            
        class_code = str(class_res.data["ma_lop"]).strip()
        class_name = class_res.data["ten_mon"]
        
        # 3. Get all students for this class from class_students junction table
        students_res = supabase.table("class_students").select("mssv, students(name)").eq("class_id", class_id).execute()
        all_students = []
        for item in (students_res.data or []):
            all_students.append({
                "mssv": item["mssv"],
                "name": item["students"]["name"] if item.get("students") else "N/A"
            })
        
        # 4. Get present students for this session
        records_res = supabase.table("attendance_records").select("mssv").eq("session_id", session_id).execute()
        present_mssvs = {r["mssv"] for r in (records_res.data or [])}
        
        # 5. Filter absent students
        absent_students = [s for s in all_students if s["mssv"] not in present_mssvs]
        
        # 6. Create Excel using openpyxl
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DanhSachVang"
        
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        if not absent_students:
            ws.append(["Thông báo"])
            ws.append([f"Buổi điểm danh ngày {now_str}: Lớp đi đủ 100%"])
        else:
            # Add Headers
            headers_list = ["STT", "MSSV", "Họ Tên", "Mã Lớp HP", "Trạng Thái", "Thời Gian"]
            ws.append(headers_list)
            
            # Add Data
            for i, student in enumerate(absent_students, 1):
                ws.append([
                    i,
                    student.get("mssv", ""),
                    student.get("name", ""),
                    class_code, # Đây là mã lớp học phần (ví dụ 225NMG03)
                    "Vắng",
                    now_str # Thời gian điểm danh
                ])
        
        # Format headers
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        wb.save(output)
        output.seek(0)
        
        # 8. Setup Response
        today_str = datetime.now().strftime("%d-%m-%Y")
        filename = f"Vang_mat_{class_code}_{today_str}.xlsx"
        
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        return StreamingResponse(
            output, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/attendance/classes/{class_id}/export-semester")
def export_semester_report(class_id: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        # 1. Get class info
        class_res = supabase.table("classes").select("ma_lop", "ten_mon").eq("id", class_id).single().execute()
        if not class_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy lớp học")
            
        class_code = class_res.data["ma_lop"]
        class_name = class_res.data["ten_mon"]

        # 2. Lấy danh sách sinh viên THỰC TẾ của lớp này (Từ bảng kết nối class_students)
        # Truy vấn join: class_students -> students
        students_res = supabase.table("class_students").select("mssv, students(name)").eq("class_id", class_id).execute()
        all_students = []
        for item in (students_res.data or []):
            all_students.append({
                "mssv": item["mssv"],
                "name": item["students"]["name"] if item.get("students") else "N/A"
            })
            
        if not all_students:
            raise HTTPException(status_code=400, detail="Lớp học này chưa có danh sách sinh viên. Vui lòng đồng bộ Excel trước.")

        # 3. Lấy TẤT CẢ dữ liệu điểm danh từ trước đến nay (giống như một file Excel tổng)
        # Bước 1: Lấy danh sách MSSV của lớp này
        mssv_list = [s["mssv"] for s in all_students]
        
        # Bước 2: Truy quét toàn bộ bảng attendance_records
        # Không cần quan tâm buổi học nào, cứ có MSSV trong danh sách lớp là lấy
        records_res = supabase.table("attendance_records").select("mssv, session_id").in_("mssv", mssv_list).execute()
        all_records = records_res.data or []
        
        # Tạo bản đồ: MSSV -> Danh sách các buổi đã tham gia (session_id duy nhất)
        attendance_map = {}
        unique_sessions = set()
        
        for r in all_records:
            m = r["mssv"]
            sid = r["session_id"]
            unique_sessions.add(sid)
            if m not in attendance_map: attendance_map[m] = set()
            attendance_map[m].add(sid)

        # 4. Tạo file Excel
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BaoCaoTongHop"
        
        # Tiêu đề báo cáo
        ws.append(["BÁO CÁO ĐIỂM DANH TỔNG HỢP (PHONG CÁCH ĐƠN GIẢN)"])
        ws.append([f"Lớp: {class_name} ({class_code})"])
        ws.append([f"Tổng số buổi đã tổ chức điểm danh: {len(unique_sessions)}"])
        ws.append([])
        
        # Header bảng
        ws.append(["STT", "MSSV", "Họ Tên", "Số buổi có mặt", "Số buổi vắng", "Ghi chú"])
        
        # Đổ dữ liệu vào
        for i, student in enumerate(all_students, 1):
            present_count = len(attendance_map.get(student["mssv"], set()))
            absent_count = len(unique_sessions) - present_count
            
            ws.append([
                i,
                student["mssv"],
                student["name"],
                present_count,
                absent_count,
                "Đạt" if absent_count <= 2 else "Vắng nhiều" # Ví dụ cảnh báo
            ])
            
        # Styling
        for cell in ws[5]: # Header row
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

        wb.save(output)
        output.seek(0)
        
        filename = f"Tong_hop_diem_danh_{class_code}.xlsx"
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        return StreamingResponse(output, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/students")
def get_all_students():
    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase is not initialized.")
    try:
        # Truy vấn sinh viên kèm theo thông tin lớp học từ bảng trung gian
        res = supabase.table("students").select("*, class_students(classes(ma_lop))").order("name").execute()
        
        # Format lại dữ liệu để tương thích với Frontend
        students = []
        for s in (res.data or []):
            class_code = None
            if s.get("class_students") and len(s["class_students"]) > 0:
                class_info = s["class_students"][0].get("classes")
                if class_info:
                    class_code = class_info.get("ma_lop")
            
            student_data = {
                "mssv": s["mssv"],
                "name": s["name"],
                "email": s.get("email"),
                "device_id": s.get("device_id"),
                "created_at": s.get("created_at"),
                "class_code": class_code
            }
            students.append(student_data)
            
        return students
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
