import openpyxl
import os

files = [
    "DanhSachSVLHP_225CHKCHST02.xlsx",
    "DanhSachSVLHP_225NMG02.xlsx",
    "DanhSachSVLHP_225NMG03.xlsx"
]

path = "d:/20260403_HocTap/quan-ly-cong-viec/danh-sach-sinh-vien/"

print("--- Excel Student Count Check ---")
for f in files:
    full_path = os.path.join(path, f)
    if not os.path.exists(full_path):
        print(f"File not found: {f}")
        continue
    
    try:
        wb = openpyxl.load_workbook(full_path, data_only=True)
        sheet = wb.active
        
        # Determine start row by looking for "Mã SV"
        start_row = 1
        for r in range(1, 15):
            for c in range(1, 10):
                val = str(sheet.cell(row=r, column=c).value or "").strip()
                if "Mã SV" in val or "MSSV" in val:
                    start_row = r + 1
                    break
            if start_row > 1: break
            
        count = 0
        for r in range(start_row, sheet.max_row + 1):
            mssv = sheet.cell(row=r, column=2).value
            if mssv and str(mssv).strip().isdigit() and len(str(mssv).strip()) >= 5:
                count += 1
        
        print(f"File: {f} | Detected Start Row: {start_row} | Student Count: {count}")
    except Exception as e:
        print(f"Error reading {f}: {e}")
